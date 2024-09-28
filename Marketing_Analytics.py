#!/usr/bin/env python
# coding: utf-8

# # Marketing Analytics for mock sample data set for customers

# We need to identify the following here basis the raw data:
# Visualize - 
# 1. Distribution of acquisition cost 
# 2. Distribution of Revenue
# 3. Customer Acquisition Cost By Channel
# 4. Conversion Rate
# 5. Total Revenue By Channel
# 6. ROAS for different channels
# 7. CLTV Distribution for different channels
#  

# In[1]:


get_ipython().system('pip install xlsxwriter')


# In[3]:


pip install plotly openpyxl kaleido


# In[1]:


pip install matplotlib openpyxl


# In[3]:


pip install -U kaleido


# In[1]:


pip install --force-reinstall kaleido==0.1.0.post1


# In[1]:


pip show openpyxl


# In[3]:


pip install --upgrade openpyxl


# In[1]:


pip show openpyxl


# In[3]:


pip install --upgrade jupyter ipython


# In[3]:


pip show jupyter


# In[1]:


import pandas as pd
import plotly.graph_objs as go
import plotly.express as px
import plotly.io as pio
# import xlsxwriter
import openpyxl
from openpyxl.drawing.image import Image # for excel file operations
import os  # Importing the os module for file path operations

from datetime import datetime  # Importing datetime to get the current date

pio.templates.default = "plotly_white"


# In[3]:


data = pd.read_csv("C:\\Users\\abhin\\Documents\\Python_Resources\\Capstone\\Customer_Lifetime_Value\\customer_acquisition_data.csv")
pd.options.display.float_format = "{:,.2f}".format


# In[5]:


print(data.head())


# In[11]:


# Open the excel file that I plan to use to save all my analytis for the data


# In[9]:


# Get the current date to create a dynamic filename
current_date = datetime.now().strftime("%Y-%m-%d")  # Format: YYYY-MM-DD
excel_file_name = f"Customer_Lifetime_Value_{current_date}.xlsx"  # Dynamic filename

excel_file_path = os.path.join("C:\\Users\\abhin\\Documents\\Python_Resources\\Capstone\\Customer_Lifetime_Value\\", excel_file_name)

# Step 2: Create a new Excel workbook or load existing one
if os.path.exists(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)  # Load existing workbook
else:
    wb = openpyxl.Workbook()  # Create a new workbook if it doesn't exist

# Step 3: Create a dynamic sheet name for the current date
sheet_name = f"Data_{current_date}"  # Create a sheet name based on the current date
print (f"sheet name: {sheet_name}")

if 'Sheet' in wb.sheetnames:
    del wb['Sheet']
    print("Sheet deleted")
else:
    print("Nothing to delete")


# Add a new sheet for the current date
if sheet_name not in wb.sheetnames:  # Check if the sheet already exists
    ws = wb.create_sheet(title=sheet_name)  # Create a new sheet
else:
    ws = wb[sheet_name]  # Use the existing sheet if it already exists

print(f"Excel file saved at: {excel_file_path} with sheet: {sheet_name}")

wb.save(excel_file_path)


# In[11]:


#I am creating a function for the repeated task of writing to the file and will pass the variable to the function

def save_to_excel(fig,image_filename,excel_cell_value):
    
    #Step 1: Save the Plotly figure as a PNG image
    image_path = image_filename  # Define the image path
    
    print(type(image_path))
    print(f"image path variable: {image_path}")
   
    try:
        # print("image writing")
        pio.write_image(fig, image_path)
        # print("image written \n")
    except Exception as e:
        print(f"error saving image: {e}")
    
    # Step 2: Add the image to the new Excel sheet
    try:
        img = Image(image_path)
        # print(type(img))
        # print(type(excel_cell_value))
        ws.add_image(img, excel_cell_value)  # Position the image at cell A1, you can make this dynamic too if there are multiple entries to a sheet as after every 12 cells etc.
        print("image added \n")
    except Exception as e:
        print(f"exception while addig image to excel: {e}")
    # Ensure the directory exists
    try:
        print(f"excel file path: {excel_file_path}")
        os.makedirs(os.path.dirname(excel_file_path), exist_ok=True)
        print("check done \n")
    except Exception as e:
        print(f"exception while ensuring directory exists: {e}")
    
    # Step 3: Save the workbook
    try:
        # print(type(excel_file_path))
        # print(type(sheet_name))
        # print(sheet_name)
        # print(excel_file_path)
        wb.save(excel_file_path)
        print(f"Excel file saved at: {excel_file_path} with sheet: {sheet_name}")
    except Exception as e:
        print(f"exception while saving the file: {e}")
    


# #### Visualizing the Distribution of Acquisition
# ***
# Here we attempt to visualize the distribution of acquisition cost for the data set shared. What I have also done is add the the image analysis to an excel sheet. 
# 
# This will help you in say generating report which can be picked by an auto job and emailed early in the morning to see the trends. 

# In[13]:


import stat

fig = px.histogram(data, 
                   x="cost ($)",
                   nbins=15, 
                   title='<b>Distribution of Acquisition Cost</b>', text_auto =True)
fig.update_layout(autosize = False, width =  900, height = 500)
fig.show()

print(f"Excel file saved at: {excel_file_path} with sheet: {sheet_name}. Saving the image file in Jupyter")
fig.write_image("C:\\Users\\abhin\\Documents\\Python_Resources\\Capstone\\Customer_Lifetime_Value\\figure.png", format="png")

#creating variables for the save to excel function parameters
image_filename= 'AcquistionCost_Histogram.png'# Define the image name
cell_value='A1' # define the cell value where the image will be pasted

#save the image to the excel
save_to_excel(fig,image_filename,cell_value)


# #### Visualizing the Revenue Distribution
# ***
# Here we attempt to visualize the revenue distribution. 

# In[15]:


fig = px.histogram(data, 
                   x="revenue ($)", 
                   nbins=20, 
                   title='<B>Distribution of Revenue  ($)</B>',text_auto=True)
fig.update_layout(autosize = False, width =  900, height = 500)
fig.show()


#creating variables for the save to excel function parameters
#Step 1: Save the Plotly figure as a PNG image
image_path = 'Revenue_Histogram.png'  # Define the image path
cell_value='P1' # define the cell value where the image will be pasted

#save the image to the excel
save_to_excel(fig,image_path,cell_value)


# In[ ]:


type(fig)


# #### Average Customer Acquistion Cost by Channel
# ***

# In[17]:


cost_by_channel = data.groupby('channel')['cost ($)'].mean().reset_index()

# channel   cost ($)  conversion_rate (%)  revenue ($)

fig = px.bar(cost_by_channel, 
             x='channel', 
             y='cost ($)', 
             title='<b>Average Customer Acquisition Cost by Channel (in $)</b>', text_auto=True)

fig.update_layout(autosize = False, width =  900, height = 500)
fig.show()

#Step 1: Save the Plotly figure as a PNG image
#creating variables for the save to excel function parameters
image_path = 'CostByChannel_Histogram.png'  # Define the image path
cell_value='A27' # define the cell value where the image will be pasted

#save the image to the excel
save_to_excel(fig,image_path,cell_value)


# #### Conversion Rate by Channel
# ***

# In[21]:


conversion_by_channel = round (data.groupby('channel')['conversion_rate (%)'].mean().reset_index(),2)

fig = px.bar(conversion_by_channel, x='channel', 
             y='conversion_rate (%)', 
             title='<b>Conversion Rate by Channel</b>', text_auto=True)
fig.update_layout(autosize = False, width =  900, height = 500)
fig.show()

#Step 1: Save the Plotly figure as a PNG image
#creating variables for the save to excel function parameters
image_path = 'ConversionRateByChannel_Histogram.png'  # Define the image path
cell_value='P27' # define the cell value where the image will be pasted

#save the image to the excel
save_to_excel(fig,image_path,cell_value)


# #### Share of Revenue by Channel
# ***

# In[29]:


revenue_by_channel = data.groupby('channel')['revenue ($)'].sum().reset_index()

fig = px.pie(revenue_by_channel, 
             values='revenue ($)', 
             names='channel', 
             title='<b>Total Revenue by Channel</b>', 
             hole=0.6, color_discrete_sequence=px.colors.qualitative.Pastel)
fig.update_traces(textposition='outside', textinfo='percent+label')
# fig.update_layout(title='Bold Title - Graphs', title_font=dict(family='Arial Black', size=24))
fig.update_layout(autosize = False, width =  900, height = 500,)
fig.show()

#Step 1: Save the Plotly figure as a PNG image
#creating variables for the save to excel function parameters
image_path = 'RevenueByChannel_Histogram.png'  # Define the image path
cell_value='A54' # define the cell value where the image will be pasted

#save the image to the excel
save_to_excel(fig,image_path,cell_value)


# #### Return on INvesmtent by Channel
# ***

# In[31]:


data['roi'] = data['revenue ($)'] / data['cost ($)']
roi_by_channel = round(data.groupby('channel')['roi'].mean().reset_index(),2)

fig = px.bar(roi_by_channel, 
             x='channel', 
             y='roi', title='<b>Return on Investment (ROI) by Channel</b>',text_auto=True)
# fig.update_layout(title='Bold Title - Graphs', title_font=dict(family='Arial Black', size=24))
fig.update_layout(autosize = False, width =  900, height = 500,)
fig.show()

#Step 1: Save the Plotly figure as a PNG image
#creating variables for the save to excel function parameters
image_path = 'ReturnOnInvestByChannel_Histogram.png'  # Define the image path
cell_value='P54' # define the cell value where the image will be pasted

#save the image to the excel
save_to_excel(fig,image_path,cell_value)


# #### CLTV
# ***
# The ROI from email marketing is way higher than all other channels, while the ROI from paid advertising is the lowest. Now let’s calculate the customer lifetime value from each channel. Based on the data we have, we can use the formula mentioned below to calculate CLTV:<br>
# 
# **CLTV = (revenue – cost) * conversion_rate / cost**st

# In[33]:


data['cltv'] = (data['revenue ($)'] - data['cost ($)']) * data['conversion_rate (%)'] / data['cost ($)']

channel_cltv = data.groupby('channel')['cltv'].mean().reset_index()

fig = px.bar(channel_cltv, x='channel', y='cltv', color='channel',
             title='<b>Customer Lifetime Value by Channel</b>')

fig.update_xaxes(title='Channel')
fig.update_yaxes(title='CLTV')
fig.update_layout(autosize = False, width =  1000, height = 500)
fig.show()

#Step 1: Save the Plotly figure as a PNG image
#creating variables for the save to excel function parameters
image_path = 'CLVByChannel_Histogram.png'  # Define the image path
cell_value='A81' # define the cell value where the image will be pasted

#save the image to the excel
save_to_excel(fig,image_path,cell_value)


# In[ ]:




